"""Singapore leave entitlement, Google Sheet staging, and payslip summaries."""

from __future__ import annotations

import json
import re
import sqlite3
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl

from payroll_engine import DATABASE


SOURCE_ID = "HR_LEAVE_SHEET"

DDL = """
CREATE TABLE IF NOT EXISTS employee_leave_entitlements (
    entitlement_id INTEGER PRIMARY KEY,
    employee_id INTEGER NOT NULL REFERENCES employees(employee_id),
    leave_year INTEGER NOT NULL,
    leave_type TEXT NOT NULL,
    entitlement_units TEXT NOT NULL CHECK(entitlement_units IN ('DAYS','HOURS')),
    entitlement_amount TEXT NOT NULL,
    carry_forward_amount TEXT NOT NULL DEFAULT '0',
    adjustment_amount TEXT NOT NULL DEFAULT '0',
    source TEXT NOT NULL,
    source_record_id TEXT,
    reviewed_at TEXT,
    UNIQUE(employee_id,leave_year,leave_type)
);

CREATE TABLE IF NOT EXISTS leave_import_records (
    leave_import_id INTEGER PRIMARY KEY,
    source_system TEXT NOT NULL,
    source_record_id TEXT NOT NULL,
    employee_id INTEGER REFERENCES employees(employee_id),
    submitted_at TEXT,
    leave_type TEXT NOT NULL,
    start_date TEXT NOT NULL,
    requested_amount TEXT NOT NULL,
    requested_units TEXT NOT NULL CHECK(requested_units IN ('DAYS','HOURS')),
    day_portion TEXT,
    reason TEXT,
    evidence_url TEXT,
    status TEXT NOT NULL DEFAULT 'PENDING' CHECK(status IN ('PENDING','APPROVED','REJECTED','CANCELLED')),
    approved_by TEXT,
    approved_at TEXT,
    source_payload TEXT NOT NULL,
    imported_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(source_system,source_record_id)
);

CREATE TABLE IF NOT EXISTS employee_leave_eligibility (
    employee_id INTEGER PRIMARY KEY REFERENCES employees(employee_id),
    youngest_child_birth_date TEXT,
    child_singapore_citizen INTEGER CHECK(child_singapore_citizen IN (0,1)),
    hospitalisation_leave_eligible INTEGER CHECK(hospitalisation_leave_eligible IN (0,1)),
    notes TEXT
);

CREATE TABLE IF NOT EXISTS employee_leave_aliases (
    alias_norm TEXT PRIMARY KEY,
    employee_id INTEGER NOT NULL REFERENCES employees(employee_id),
    source TEXT NOT NULL,
    reviewed_at TEXT
);
"""


def _norm(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def _leave_type(raw: object) -> str:
    value = _norm(raw)
    if value in {"annual leave"}: return "ANNUAL_LEAVE"
    if value in {"medical leave", "mc"}: return "SICK_LEAVE"
    if value in {"child care", "child care leave", "childcare leave", "ccl"}: return "CHILDCARE_LEAVE"
    if value in {"unpaid leave"}: return "UNPAID_LEAVE"
    if value in {"fcl", "family care leave"}: return "FAMILY_CARE_LEAVE"
    if value in {"compassionate leave"}: return "COMPASSIONATE_LEAVE"
    return "OTHER_LEAVE"


def migrate(connection: sqlite3.Connection) -> None:
    connection.executescript(DDL)
    connection.execute("UPDATE leave_policies SET applies_to='ALL_EMPLOYEES' WHERE leave_type='ANNUAL_LEAVE'")
    connection.execute(
        """INSERT INTO application_scope(scope_key,scope_value) VALUES('PART_TIME_LEAVE','MOM_PRORATED_HOURS')
           ON CONFLICT(scope_key) DO UPDATE SET scope_value=excluded.scope_value,updated_at=CURRENT_TIMESTAMP"""
    )


def resolve_employee(connection: sqlite3.Connection, source_name: object) -> int | None:
    alias = _norm(source_name)
    stored = connection.execute(
        "SELECT employee_id FROM employee_leave_aliases WHERE alias_norm=?", (alias,)
    ).fetchone()
    if stored:
        return int(stored[0])
    matches = connection.execute(
        "SELECT employee_id FROM employees WHERE name_norm LIKE ?", (f"%{alias}%",)
    ).fetchall()
    return int(matches[0][0]) if len(matches) == 1 else None


def import_google_export(connection: sqlite3.Connection, workbook_path: Path) -> dict[str, int]:
    """Synchronise 2026 HR-confirmed rows and import explicit annual entitlements."""
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    staged = unmatched = entitlements = 0
    connection.execute(
        """UPDATE leave_import_records SET status='CANCELLED'
           WHERE source_system='GOOGLE_SHEET' AND substr(start_date,1,4)='2026'"""
    )
    form = workbook["Form responses 1"]
    for row_number in range(2, form.max_row + 1):
        values = [form.cell(row_number, column).value for column in range(1, 17)]
        start = values[6]
        if not isinstance(start, datetime) or start.year != 2026:
            continue
        employee_id = resolve_employee(connection, values[9])
        if employee_id is None:
            unmatched += 1
        submitted = values[0].isoformat() if isinstance(values[0], datetime) else None
        source_id = f"{SOURCE_ID}:Form responses 1:{row_number}"
        payload = json.dumps({str(index + 1): value.isoformat() if isinstance(value, datetime) else value
                              for index, value in enumerate(values)}, ensure_ascii=False, default=str)
        connection.execute(
            """INSERT INTO leave_import_records(source_system,source_record_id,employee_id,submitted_at,
               leave_type,start_date,requested_amount,requested_units,day_portion,reason,evidence_url,
               source_payload,status,approved_by,approved_at)
               VALUES('GOOGLE_SHEET',?,?,?,?,?,?,?,?,?,?,?,'APPROVED','HR Google Sheet',CURRENT_TIMESTAMP)
               ON CONFLICT(source_system,source_record_id) DO UPDATE SET employee_id=excluded.employee_id,
               submitted_at=excluded.submitted_at,leave_type=excluded.leave_type,start_date=excluded.start_date,
               requested_amount=excluded.requested_amount,day_portion=excluded.day_portion,
               reason=excluded.reason,evidence_url=excluded.evidence_url,source_payload=excluded.source_payload,
               status='APPROVED',approved_by='HR Google Sheet',approved_at=CURRENT_TIMESTAMP,
               imported_at=CURRENT_TIMESTAMP""",
            (source_id, employee_id, submitted, _leave_type(values[3]), start.date().isoformat(),
             str(Decimal(str(values[7] or 0))), "DAYS", str(values[5] or ""), values[8], values[12], payload),
        )
        staged += 1

    entitlement_sheet = workbook["Leave entitlement"]
    mapping = {"AL": "ANNUAL_LEAVE", "MC": "SICK_LEAVE", "CCL": "CHILDCARE_LEAVE", "OTH": "OTHER_LEAVE"}
    headers = [entitlement_sheet.cell(3, column).value for column in range(3, 7)]
    for row_number in range(4, 11):
        employee_id = resolve_employee(connection, entitlement_sheet.cell(row_number, 2).value)
        if not employee_id:
            continue
        for offset, header in enumerate(headers, start=3):
            amount = entitlement_sheet.cell(row_number, offset).value
            if amount is None:
                continue
            connection.execute(
                """INSERT INTO employee_leave_entitlements(employee_id,leave_year,leave_type,
                   entitlement_units,entitlement_amount,source,source_record_id)
                   VALUES(?,2026,?,'DAYS',?,'Google Sheet 2026 entitlement table',?)
                   ON CONFLICT(employee_id,leave_year,leave_type) DO UPDATE SET
                   entitlement_amount=excluded.entitlement_amount,source=excluded.source,
                   source_record_id=excluded.source_record_id""",
                (employee_id, mapping.get(str(header), "OTHER_LEAVE"), str(Decimal(str(amount))),
                 f"{SOURCE_ID}:Leave entitlement:{row_number}:{offset}"),
            )
            entitlements += 1
    return {"staged": staged, "unmatched": unmatched, "entitlements": entitlements}


def sick_leave_entitlement(months_of_service: int) -> tuple[int, int]:
    """MOM outpatient and inclusive hospitalisation days after 3 months of service."""
    if months_of_service < 3: return 0, 0
    if months_of_service == 3: return 5, 15
    if months_of_service == 4: return 8, 30
    if months_of_service == 5: return 11, 45
    return 14, 60


def company_annual_days(service_year: int) -> int:
    """Company policy is more generous than MOM: 11 days plus one, capped at 14."""
    return min(11 + max(service_year - 1, 0), 14)


def leave_summary(connection: sqlite3.Connection, employee_id: int, as_of: date) -> list[dict[str, object]]:
    rows = connection.execute(
        """SELECT leave_type,entitlement_units,entitlement_amount,carry_forward_amount,adjustment_amount
             FROM employee_leave_entitlements WHERE employee_id=? AND leave_year=?
             ORDER BY CASE leave_type WHEN 'ANNUAL_LEAVE' THEN 1 WHEN 'SICK_LEAVE' THEN 2
                      WHEN 'CHILDCARE_LEAVE' THEN 3 ELSE 4 END""", (employee_id, as_of.year),
    ).fetchall()
    result = []
    for row in rows:
        used = connection.execute(
            """SELECT COALESCE(SUM(CAST(requested_amount AS REAL)),0) FROM leave_import_records
               WHERE employee_id=? AND leave_type=? AND status='APPROVED'
                 AND substr(start_date,1,4)=? AND start_date<=?""",
            (employee_id, row[0], str(as_of.year), as_of.isoformat()),
        ).fetchone()[0]
        entitlement = Decimal(row[2]) + Decimal(row[3]) + Decimal(row[4])
        used_decimal = Decimal(str(used))
        result.append({"leave_type": row[0], "units": row[1], "entitlement": entitlement,
                       "used": used_decimal, "remaining": entitlement - used_decimal})
    return result


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", nargs="?", type=Path)
    args = parser.parse_args()
    connection = sqlite3.connect(DATABASE)
    with connection:
        migrate(connection)
        print(import_google_export(connection, args.workbook) if args.workbook else "Leave schema installed")
