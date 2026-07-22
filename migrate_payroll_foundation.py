"""Add payroll-domain tables and import a conservative June 2026 baseline."""

from __future__ import annotations

import sqlite3
import os
from datetime import datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook

PROJECT = Path(__file__).parent
DATABASE = PROJECT / "data" / "employees.db"
WORKBOOK = Path(os.environ.get("PAYROLL_SOURCE_WORKBOOK", "payroll-source.xlsx"))
EFFECTIVE_FROM = "2026-06-01"


def cents(value) -> int | None:
    if value in (None, ""):
        return None
    return int((Decimal(str(value)) * 100).quantize(Decimal("1"), ROUND_HALF_UP))


def is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=") or hasattr(value, "text")


connection = sqlite3.connect(DATABASE)
connection.row_factory = sqlite3.Row
connection.execute("PRAGMA foreign_keys = ON")

try:
    connection.execute("BEGIN")
    connection.executescript(
        """
        CREATE TABLE IF NOT EXISTS employment_terms (
            employment_term_id INTEGER PRIMARY KEY,
            employee_id INTEGER NOT NULL REFERENCES employees(employee_id),
            effective_from TEXT NOT NULL,
            effective_to TEXT,
            pay_basis TEXT CHECK(pay_basis IN ('MONTHLY', 'HOURLY')),
            monthly_salary_cents INTEGER CHECK(monthly_salary_cents >= 0),
            hourly_rate_cents INTEGER CHECK(hourly_rate_cents >= 0),
            standard_weekly_minutes INTEGER CHECK(standard_weekly_minutes >= 0),
            overtime_eligible INTEGER CHECK(overtime_eligible IN (0, 1)),
            overtime_multiplier TEXT,
            branch_code TEXT,
            source TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(employee_id, effective_from)
        );

        CREATE TABLE IF NOT EXISTS compliance_profiles (
            compliance_profile_id INTEGER PRIMARY KEY,
            employee_id INTEGER NOT NULL REFERENCES employees(employee_id),
            effective_from TEXT NOT NULL,
            effective_to TEXT,
            residency_status TEXT NOT NULL CHECK(
                residency_status IN ('CITIZEN', 'PR', 'FOREIGNER', 'CITIZEN_OR_PR_REVIEW')
            ),
            work_pass_type TEXT,
            pass_start_date TEXT,
            pass_expiry_date TEXT,
            pr_effective_date TEXT,
            pr_contribution_scheme TEXT,
            cpf_applicable INTEGER CHECK(cpf_applicable IN (0, 1)),
            shg_fund TEXT,
            shg_override_cents INTEGER,
            levy_skill_level TEXT,
            levy_tier TEXT,
            actual_monthly_levy_cents INTEGER,
            tax_clearance_required INTEGER CHECK(tax_clearance_required IN (0, 1)),
            source TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(employee_id, effective_from)
        );

        CREATE TABLE IF NOT EXISTS payroll_runs (
            payroll_run_id INTEGER PRIMARY KEY,
            period_start TEXT NOT NULL,
            period_end TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'DRAFT' CHECK(
                status IN ('DRAFT', 'REVIEW', 'APPROVED', 'LOCKED', 'VOID')
            ),
            calculation_version TEXT NOT NULL,
            notes TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            approved_at TEXT,
            locked_at TEXT,
            UNIQUE(period_start, period_end)
        );

        CREATE TABLE IF NOT EXISTS attendance_imports (
            attendance_import_id INTEGER PRIMARY KEY,
            source_name TEXT NOT NULL,
            source_hash TEXT NOT NULL UNIQUE,
            period_start TEXT NOT NULL,
            period_end TEXT NOT NULL,
            imported_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS attendance_entries (
            attendance_entry_id INTEGER PRIMARY KEY,
            attendance_import_id INTEGER NOT NULL REFERENCES attendance_imports(attendance_import_id),
            employee_id INTEGER REFERENCES employees(employee_id),
            source_employee_name TEXT NOT NULL,
            work_date TEXT NOT NULL,
            worked_minutes INTEGER NOT NULL CHECK(worked_minutes >= 0),
            UNIQUE(attendance_import_id, source_employee_name, work_date)
        );

        CREATE TABLE IF NOT EXISTS attendance_adjustments (
            attendance_adjustment_id INTEGER PRIMARY KEY,
            employee_id INTEGER NOT NULL REFERENCES employees(employee_id),
            work_date TEXT NOT NULL,
            adjustment_minutes INTEGER NOT NULL,
            reason TEXT NOT NULL,
            created_by TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            approved_at TEXT
        );

        CREATE TABLE IF NOT EXISTS legacy_payroll_baseline (
            baseline_id INTEGER PRIMARY KEY,
            employee_id INTEGER NOT NULL REFERENCES employees(employee_id),
            period_start TEXT NOT NULL,
            employee_code TEXT,
            monthly_wage_cents INTEGER,
            hourly_rate_cents INTEGER,
            worked_minutes INTEGER,
            gross_pay_cents INTEGER,
            employer_cpf_cents INTEGER,
            employee_cpf_cents INTEGER,
            net_pay_cents INTEGER,
            source TEXT NOT NULL,
            UNIQUE(employee_id, period_start)
        );
        """
    )

    # Seed effective-dated compliance from fields the user has already reviewed.
    employees = connection.execute(
        """SELECT employee_id, visa_status, main_branch, association,
                  association_fee FROM employees"""
    ).fetchall()
    for employee in employees:
        if employee["visa_status"] == "WP":
            residency, pass_type, cpf = "FOREIGNER", "WP", 0
        else:
            residency, pass_type, cpf = "CITIZEN_OR_PR_REVIEW", None, 1
        connection.execute(
            """INSERT INTO compliance_profiles
               (employee_id, effective_from, residency_status, work_pass_type,
                cpf_applicable, shg_fund, shg_override_cents, source)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)
               ON CONFLICT(employee_id, effective_from) DO UPDATE SET
                 residency_status=excluded.residency_status,
                 work_pass_type=excluded.work_pass_type,
                 cpf_applicable=excluded.cpf_applicable,
                 shg_fund=excluded.shg_fund,
                 shg_override_cents=excluded.shg_override_cents,
                 source=excluded.source""",
            (
                employee["employee_id"], EFFECTIVE_FROM, residency, pass_type, cpf,
                employee["association"], cents(employee["association_fee"]),
                "reviewed visa_status and legacy SHG migration",
            ),
        )

    formula_book = load_workbook(WORKBOOK, data_only=False, read_only=False)
    value_book = load_workbook(WORKBOOK, data_only=True, read_only=False)
    formulas = formula_book["June 2026"]
    values = value_book["June 2026"]

    matched = baselines = terms = skipped_formula_salary = 0
    for row in range(3, 28):
        name = str(values.cell(row, 1).value or "").strip()
        if not name:
            continue
        employee = connection.execute(
            "SELECT employee_id, main_branch FROM employees WHERE name_norm = ?",
            (" ".join(name.casefold().split()),),
        ).fetchone()
        if employee is None:
            continue
        matched += 1
        pay_basis = "HOURLY" if str(values.cell(row, 10).value or "").strip() == "PT" else "MONTHLY"
        monthly_value = values.cell(row, 11).value
        hourly_value = values.cell(row, 12).value
        connection.execute(
            """INSERT INTO legacy_payroll_baseline
               (employee_id, period_start, employee_code, monthly_wage_cents,
                hourly_rate_cents, worked_minutes, gross_pay_cents,
                employer_cpf_cents, employee_cpf_cents, net_pay_cents, source)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
               ON CONFLICT(employee_id, period_start) DO UPDATE SET
                 employee_code=excluded.employee_code,
                 monthly_wage_cents=excluded.monthly_wage_cents,
                 hourly_rate_cents=excluded.hourly_rate_cents,
                 worked_minutes=excluded.worked_minutes,
                 gross_pay_cents=excluded.gross_pay_cents,
                 employer_cpf_cents=excluded.employer_cpf_cents,
                 employee_cpf_cents=excluded.employee_cpf_cents,
                 net_pay_cents=excluded.net_pay_cents,
                 source=excluded.source""",
            (
                employee["employee_id"], EFFECTIVE_FROM,
                str(values.cell(row, 10).value or "").strip() or None,
                cents(monthly_value), cents(hourly_value),
                int(Decimal(str(values.cell(row, 14).value or 0)) * 60),
                cents(values.cell(row, 17).value), cents(values.cell(row, 18).value),
                cents(values.cell(row, 19).value), cents(values.cell(row, 20).value),
                "Full Time Salary and CPF.xlsx / June 2026",
            ),
        )
        baselines += 1

        contractual_monthly = None
        contractual_hourly = None
        if pay_basis == "HOURLY" and hourly_value not in (None, "") and not is_formula(formulas.cell(row, 12).value):
            contractual_hourly = cents(hourly_value)
        elif pay_basis == "MONTHLY" and monthly_value not in (None, ""):
            if is_formula(formulas.cell(row, 11).value):
                skipped_formula_salary += 1
            else:
                contractual_monthly = cents(monthly_value)
        if contractual_monthly is not None or contractual_hourly is not None:
            connection.execute(
                """INSERT INTO employment_terms
                   (employee_id, effective_from, pay_basis, monthly_salary_cents,
                    hourly_rate_cents, standard_weekly_minutes, branch_code, source)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                   ON CONFLICT(employee_id, effective_from) DO UPDATE SET
                     pay_basis=excluded.pay_basis,
                     monthly_salary_cents=excluded.monthly_salary_cents,
                     hourly_rate_cents=excluded.hourly_rate_cents,
                     standard_weekly_minutes=excluded.standard_weekly_minutes,
                     branch_code=excluded.branch_code,
                     source=excluded.source""",
                (
                    employee["employee_id"], EFFECTIVE_FROM, pay_basis,
                    contractual_monthly, contractual_hourly,
                    44 * 60 if pay_basis == "MONTHLY" else None,
                    employee["main_branch"], "unambiguous June 2026 workbook value",
                ),
            )
            terms += 1

    connection.commit()
    print(f"COMPLIANCE_PROFILES={len(employees)}")
    print(f"JUNE_NAMES_MATCHED={matched}")
    print(f"JUNE_BASELINES={baselines}")
    print(f"EMPLOYMENT_TERMS_IMPORTED={terms}")
    print(f"FORMULA_DERIVED_SALARIES_SKIPPED={skipped_formula_salary}")
finally:
    connection.close()
